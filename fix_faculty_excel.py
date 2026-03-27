import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Faculty"

# Define headers
headers = ["Name", "Email", "Mobile", "Department", "Profession", "Valid Till"]

# Add header row with formatting
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font

# Sample data for faculty
sample_data = [
    ["Dr. Rajesh Gupta", "rajesh.gupta@pict.edu", "9988776655", "Computer Science", "Professor", "2027-12-31"],
    ["Dr. Priya Sharma", "priya.sharma@pict.edu", "9988776656", "Electronics", "Associate Professor", "2027-12-31"],
    ["Dr. Amit Patel", "amit.patel@pict.edu", "9988776657", "Mechanical", "Assistant Professor", "2027-12-31"],
    ["Ms. Neha Singh", "neha.singh@pict.edu", "9988776658", "Civil", "Lecturer", "2027-12-31"],
    ["Dr. Vikram Desai", "vikram.desai@pict.edu", "9988776659", "Electrical", "Professor", "2027-12-31"],
    ["Ms. Anjali Kulkarni", "anjali.kulkarni@pict.edu", "9988776660", "Computer Science", "Lecturer", "2027-12-31"],
    ["Dr. Suresh Nair", "suresh.nair@pict.edu", "9988776661", "Information Technology", "Assistant Professor", "2027-12-31"],
    ["Ms. Divya Sharma", "divya.sharma@pict.edu", "9988776662", "Electronics", "Lecturer", "2027-12-31"],
]

# Add data rows
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

# Adjust column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 15

# Save the file
output_path = "sample_excel_files/faculty_bulk_upload.xlsx"
wb.save(output_path)

print(f"✅ Faculty Excel file created: {output_path}")
print("\nFormat:")
print("Column 1: Name (Text)")
print("Column 2: Email (Email format)")
print("Column 3: Mobile (10-digit number)")
print("Column 4: Department (Text)")
print("Column 5: Profession (Text)")
print("Column 6: Valid Till (Date in YYYY-MM-DD format)")
print(f"\n✅ Sample data: {len(sample_data)} faculty members added")
