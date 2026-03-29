import openpyxl
from openpyxl.styles import Font, PatternFill

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Students"

# Define headers
# Note: RegNo/PRN is provided from the institution, not auto-generated
headers = ["RegNo/PRN", "Name", "Email", "Mobile", "DOB", "Year", "Department"]

# Add header row with formatting
header_fill = PatternFill(start_color="0e7490", end_color="0e7490", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font

# Add instruction row
instruction_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
instruction_font = Font(color="0c4a6e", bold=False, size=9)
instruction_cell = ws.cell(row=2, column=1)
instruction_cell.value = "IMPORTANT: Use your institution's PRN/Reg No. Each student has a unique permanent registration number. Department is optional."
instruction_cell.fill = instruction_fill
instruction_cell.font = instruction_font
ws.merge_cells('A2:G2')

# Sample data for students
sample_data = [
    ["12101001", "Rahul Sharma", "rahul.sharma@pict.edu", "9876543210", "2003-03-15", 3, "Computer Engineering"],
    ["12101002", "Priya Singh", "priya.singh@pict.edu", "9876543211", "2003-06-20", 3, "Information Technology"],
    ["12101003", "Arjun Patel", "arjun.patel@pict.edu", "9876543212", "2003-01-10", 3, "Electronics Engineering"],
    ["12101004", "Neha Gupta", "neha.gupta@pict.edu", "9876543213", "2004-05-25", 2, "Mechanical Engineering"],
    ["12101005", "Vikram Desai", "vikram.desai@pict.edu", "9876543214", "2004-08-12", 2, "Civil Engineering"],
]

# Add data rows
for row_idx, row_data in enumerate(sample_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 25

# Add notes at bottom
note_row = len(sample_data) + 4
notes = [
    "FORMAT GUIDE:",
    "• RegNo/PRN: Unique permanent registration number from your institution",
    "• Name: Student name (Text)",
    "• Email: Student email address (must be valid email format)",
    "• Mobile: 10-digit mobile number without spaces or dashes",
    "• DOB: Date of birth in YYYY-MM-DD format (e.g., 2003-03-15)",
    "• Year: Current year of study (1, 2, 3, or 4)",
    "",
    "IMPORTANT NOTES:",
    "✓ Each student's PRN/RegNo must be UNIQUE",
    "✓ System uses your institution's registration numbers (NOT auto-generated)",
    "✓ Duplicate RegNo, email, or mobile numbers will be rejected",
    "✓ All fields are REQUIRED",
]

note_font = Font(color="0c4a6e", size=9, italic=True)
for offset, note in enumerate(notes):
    cell = ws.cell(row=note_row + offset, column=1)
    cell.value = note
    cell.font = note_font

# Save the file
output_path = "sample_excel_files/students_bulk_upload.xlsx"
wb.save(output_path)

print(f"✅ Student Excel file created: {output_path}")
print("\nFormat Guide:")
print("Column 1: RegNo/PRN (Unique registration number from institution)")
print("Column 2: Name (Student name - Text)")
print("Column 3: Email (Email address)")
print("Column 4: Mobile (10-digit mobile number)")
print("Column 5: DOB (Date of birth - YYYY-MM-DD)")
print("Column 6: Year (Current year: 1, 2, 3, or 4)")
print("Column 7: Department (e.g., COMPS, IT, EXTC, ETRX, MECH, CIVIL)")
print("\nNOTE: Each student's PRN/RegNo is UNIQUE and comes from your institution")
print(f"\n✅ Sample data: {len(sample_data)} students added")
