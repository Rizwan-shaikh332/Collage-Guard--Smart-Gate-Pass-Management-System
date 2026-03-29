import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

# Create a new workbook
wb = openpyxl.Workbook()

# ===== Sheet 1: PICT Students =====
ws1 = wb.active
ws1.title = "PICT Students"

# Define headers for PICT students
headers_pict = ["Reg No", "Notes"]

# Add header row with formatting
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col, header in enumerate(headers_pict, 1):
    cell = ws1.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add instruction row
instruction_cell = ws1.cell(row=2, column=1)
instruction_cell.value = "Instructions: Add registration numbers (GRN) of PICT students here"
instruction_cell.font = Font(italic=True, color="C00000", size=10)

# Add separator
ws1.cell(row=3, column=1).value = ""

# Add example data for registered students (leave RegNo empty, system will fetch automatically)
header_row = 4
headers_pict_data = ["Reg No", "Name (Auto-filled)", "Email (Auto-filled)"]
header_fill_2 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
for col, header in enumerate(headers_pict_data, 1):
    cell = ws1.cell(row=header_row, column=col)
    cell.value = header
    cell.fill = header_fill_2
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample PICT student data
sample_pict_data = [
    ["12101001"],
    ["12101002"],
    ["12101003"],
    ["12101004"],
    ["12101005"],
]

row_num = header_row + 1
for row_data in sample_pict_data:
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_num, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")
    row_num += 1

# Set column widths for PICT sheet
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 15

# ===== Sheet 2: External Students =====
ws2 = wb.create_sheet("External Students")

# Define headers for external students
headers_external = ["Full Name", "Email", "Mobile"]

for col, header in enumerate(headers_external, 1):
    cell = ws2.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add instruction row
instruction_cell_2 = ws2.cell(row=2, column=1)
instruction_cell_2.value = "⚠️ External ID auto-generated as: EventName_exe001, exe002, etc. | Fill columns below only"
instruction_cell_2.font = Font(italic=True, color="C00000", size=10)

# Add notes about duplicates and QR codes
note_cell = ws2.cell(row=3, column=1)
note_cell.value = "✓ Duplicates skipped (checked by email/mobile) | ✓ QR codes auto-emailed"
note_cell.font = Font(italic=True, color="00B050", size=9)

# Add separator
ws2.cell(row=4, column=1).value = ""

# Add example data headers
header_row_ext = 5
for col, header in enumerate(headers_external, 1):
    cell = ws2.cell(row=header_row_ext, column=col)
    cell.value = header
    cell.fill = header_fill_2
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample external student data
sample_ext_data = [
    ["Raj Singh", "raj.singh@gmail.com", "9876543210"],
    ["Priya Patel", "priya.p@gmail.com", "9876543211"],
    ["Amit Kumar", "amit.k@example.com", "9876543212"],
    ["Neha Sharma", "neha.sharma@email.com", "9876543213"],
    ["Vikram Desai", "v.desai@gmail.com", "9876543214"],
]

row_num_ext = header_row_ext + 1
for row_data in sample_ext_data:
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_num_ext, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row_num_ext += 1

# Set column widths for external sheet
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 15

# ===== Sheet 3: Help & Instructions =====
ws3 = wb.create_sheet("Help & Instructions")

instructions_data = [
    ("TECH FEST 2026 - EVENT STUDENT UPLOAD", Font(bold=True, size=14, color="FFFFFF"), PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")),
    ("", Font(size=10), None),
    ("📋 UPLOADING STUDENTS", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("✓ Use 'PICT Students' sheet for PICT registered students (only Reg No needed)", Font(size=10), None),
    ("✓ Use 'External Students' sheet for students from other colleges (3 columns only)", Font(size=10), None),
    ("✓ You can upload them separately or mix both types", Font(size=10), None),
    ("", Font(size=10), None),
    ("🚀 EXTERNAL STUDENT FORMAT (Simple & Clear)", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("Column 1: Full Name", Font(size=10), None),
    ("Column 2: Email Address", Font(size=10), None),
    ("Column 3: Mobile Number (10 digits)", Font(size=10), None),
    ("", Font(size=10), None),
    ("✅ External ID auto-generated as:", Font(size=10, bold=True), None),
    ("   Format: EventName_exe001, EventName_exe002, etc.", Font(size=10, italic=True, color="C00000"), None),
    ("   Example: Tech_Fest_2026_exe001", Font(size=10, italic=True, color="C00000"), None),
    ("", Font(size=10), None),
    ("✓ DUPLICATE CHECK", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("✓ Same student won't be added twice (checked by email & mobile)", Font(size=10), None),
    ("✓ If duplicate found: You get 'skipped' message with existing Student ID", Font(size=10), None),
    ("", Font(size=10), None),
    ("📧 QR CODE EMAILS", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("✓ All external students auto-receive QR code email after upload", Font(size=10), None),
    ("✓ Email = Main delivery method for QR code passes", Font(size=10), None),
    ("✓ Can resend anytime using 'Resend Email' button in Admin Events tab", Font(size=10), None),
    ("", Font(size=10), None),
    ("⚠️ REQUIRED FIELDS FOR EXTERNAL STUDENT", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("✓ Full Name: Complete name of student", Font(size=10), None),
    ("✓ Email: Valid email (MUST be correct - QR sent here)", Font(size=10), None),
    ("✓ Mobile: 10-digit phone number", Font(size=10), None),
    ("", Font(size=10), None),
    ("❌ NO NEED TO PROVIDE", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("✗ External ID: Auto-generated by system", Font(size=10), None),
    ("✗ Any other identifier: System handles registration", Font(size=10), None),
    ("", Font(size=10), None),
    ("📊 PICT STUDENT FORMAT", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("✓ Only column 1: PICT Registration Number (Reg No/GRN)", Font(size=10), None),
    ("✓ Name, Email, Mobile auto-filled from PICT system", Font(size=10), None),
    ("✓ Example: 12101001", Font(size=10), None),
    ("", Font(size=10), None),
    ("⚡ BULK UPLOAD TIPS", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("✓ Prepare data in correct Excel sheet (PICT vs External)", Font(size=10), None),
    ("✓ No empty rows in middle of data", Font(size=10), None),
    ("✓ Double-check email addresses (QR codes sent there)", Font(size=10), None),
    ("✓ For external students: Name | Email | Mobile (3 columns only)", Font(size=10), None),
    ("", Font(size=10), None),
    ("🎯 WORKFLOW EXAMPLE - Adding External Students", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("1. Admin Dashboard → Events Tab → Select Event (Tech Fest 2026)", Font(size=10), None),
    ("2. Click 'View Students' → External Student tab", Font(size=10), None),
    ("3. Enter: Name | Email | Mobile (system generates ID automatically)", Font(size=10), None),
    ("4. Click 'Add & Send QR Email' → Student gets QR code via email ✓", Font(size=10), None),
    ("", Font(size=10), None),
    ("🎯 BULK UPLOAD EXAMPLE - Excel Format", Font(bold=True, size=11, color="1F4E78"), None),
    ("", Font(size=10), None),
    ("Sheet: 'External Students'", Font(size=10, bold=True), None),
    ("Row 1: Full Name | Email | Mobile", Font(size=10, italic=True), None),
    ("Row 2: Raj Singh | raj.singh@gmail.com | 9876543210", Font(size=10, italic=True, color="00B050"), None),
    ("Row 3: Priya Patel | priya.p@gmail.com | 9876543211", Font(size=10, italic=True, color="00B050"), None),
    ("", Font(size=10), None),
    ("Result: Both students get Tech_Fest_2026_exe001 & exe002 IDs with QR emails ✓", Font(size=10, color="00B050", bold=True), None),
]

for row_idx, (text, font, fill) in enumerate(instructions_data, 1):
    cell = ws3.cell(row=row_idx, column=1)
    cell.value = text
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws3.row_dimensions[row_idx].height = 20

ws3.column_dimensions['A'].width = 70

# Save the file
output_path = "sample_excel_files/event_students_bulk_upload.xlsx"
wb.save(output_path)

print(f"✅ Event Students Excel file created: {output_path}")
print("\n📊 Three sheets created:")
print("   1. PICT Students - For registered PICT students")
print("   2. External Students - For students from other colleges")
print("   3. Help & Instructions - Complete guide")
print("\n🎯 Features:")
print("   ✓ Auto-generated External IDs (EventName_exe001 format)")
print("   ✓ Duplicate checking by email/mobile")
print("   ✓ QR code automatic email sending")
print("   ✓ Resend email functionality available")
print("   ✓ Bulk upload with error reporting")

