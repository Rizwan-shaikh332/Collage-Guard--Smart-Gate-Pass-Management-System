from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Depends
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import qrcode
import io
import base64
import openpyxl
import resend
import asyncio
from passlib.context import CryptContext

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend setup
resend.api_key = os.getenv('RESEND_API_KEY', '')
SENDER_EMAIL = os.getenv('SENDER_EMAIL', 'onboarding@resend.dev')

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============ MODELS ============

class Student(BaseModel):
    model_config = ConfigDict(extra="ignore")
    reg_no: str
    name: str
    email: EmailStr
    mobile_no: str
    dob: str
    current_year: int
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Faculty(BaseModel):
    model_config = ConfigDict(extra="ignore")
    faculty_id: str
    name: str
    email: EmailStr
    mobile_no: str
    department: str
    profession: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Visitor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    name: str
    email: EmailStr
    mobile_no: str
    person_to_visit: str
    photo_base64: Optional[str] = None
    purpose: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str = Field(default_factory=lambda: (datetime.now(timezone.utc) + timedelta(days=1)).isoformat())
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Alumni(BaseModel):
    model_config = ConfigDict(extra="ignore")
    name: str
    email: EmailStr
    mobile_no: str
    college_department: str
    photo_base64: Optional[str] = None
    purpose: str
    current_company: str
    job_position: str
    specialization: str
    ctc_monthly: str
    other_details: Optional[str] = None
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str = Field(default_factory=lambda: (datetime.now(timezone.utc) + timedelta(days=1)).isoformat())
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Event(BaseModel):
    model_config = ConfigDict(extra="ignore")
    event_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_name: str
    event_type: str
    date_from: str
    date_to: str
    description: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class EventStudent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    event_id: str
    reg_no: str
    name: str
    email: EmailStr
    mobile_no: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_from: str
    valid_to: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Request models
class AdminLogin(BaseModel):
    username: str
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    mobile_no: str

class GuardLogin(BaseModel):
    username: str
    password: str

class StudentCreate(BaseModel):
    reg_no: str
    name: str
    email: EmailStr
    mobile_no: str
    dob: str
    current_year: int

class FacultyCreate(BaseModel):
    name: str
    email: EmailStr
    mobile_no: str
    department: str
    profession: str
    valid_till: str

class EventCreate(BaseModel):
    event_name: str
    event_type: str
    date_from: str
    date_to: str
    description: str

class EventStudentCreate(BaseModel):
    event_id: str
    reg_no: str

class QRValidate(BaseModel):
    token: str

# ============ HELPER FUNCTIONS ============

def calculate_student_validity(current_year: int) -> str:
    """Calculate validity based on current year"""
    today = datetime.now(timezone.utc)
    years_remaining = 4 - current_year
    if years_remaining <= 0:
        return (today + timedelta(days=365)).isoformat()
    valid_till = today + timedelta(days=365 * years_remaining)
    return valid_till.isoformat()

async def get_next_faculty_id() -> str:
    """Generate next faculty ID in format fact_0001"""
    last_faculty = await db.faculty.find_one(sort=[("created_at", -1)], projection={"_id": 0, "faculty_id": 1})
    if not last_faculty:
        return "fact_0001"
    last_id = last_faculty.get("faculty_id", "fact_0000")
    num = int(last_id.split("_")[1]) + 1
    return f"fact_{num:04d}"

def generate_qr_code(token: str) -> str:
    """Generate QR code as base64 string"""
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(token)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode()

async def send_qr_email(recipient_email: str, name: str, token: str, valid_till: str):
    """Send QR code via email"""
    if not resend.api_key:
        logger.warning("Resend API key not configured")
        return
    
    qr_base64 = generate_qr_code(token)
    
    html_content = f"""
    <html>
        <body style="font-family: Arial, sans-serif; padding: 20px;">
            <h2>PICT Guard - Your Gate Pass</h2>
            <p>Hello {name},</p>
            <p>Your gate pass has been generated successfully.</p>
            <p><strong>Valid Until:</strong> {valid_till}</p>
            <div style="text-align: center; margin: 30px 0;">
                <img src="data:image/png;base64,{qr_base64}" alt="QR Code" style="max-width: 300px;"/>
            </div>
            <p>Please show this QR code at the gate for entry.</p>
            <p>Best regards,<br>PICT Guard Team</p>
        </body>
    </html>
    """
    
    params = {
        "from": SENDER_EMAIL,
        "to": [recipient_email],
        "subject": "PICT Guard - Your Gate Pass QR Code",
        "html": html_content
    }
    
    try:
        await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email sent to {recipient_email}")
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")

# ============ AUTH ROUTES ============

@api_router.post("/auth/admin/login")
async def admin_login(data: AdminLogin):
    if data.username == "admin" and data.password == "admin123":
        return {"success": True, "role": "admin", "token": str(uuid.uuid4())}
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.post("/auth/user/login")
async def user_login(data: UserLogin):
    # Check student
    student = await db.students.find_one(
        {"email": data.email, "mobile_no": data.mobile_no},
        {"_id": 0}
    )
    if student:
        return {"success": True, "role": "student", "data": student}
    
    # Check faculty
    faculty = await db.faculty.find_one(
        {"email": data.email, "mobile_no": data.mobile_no},
        {"_id": 0}
    )
    if faculty:
        return {"success": True, "role": "faculty", "data": faculty}
    
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.post("/auth/guard/login")
async def guard_login(data: GuardLogin):
    # Simple guard login - username can be guard1, guard2, etc., password is "guard123"
    if data.username.startswith("guard") and data.password == "guard123":
        return {"success": True, "role": "guard", "guard_id": data.username}
    raise HTTPException(status_code=401, detail="Invalid credentials")

# ============ STUDENT ROUTES ============

@api_router.post("/students", response_model=Student)
async def create_student(data: StudentCreate):
    # Check if student exists
    existing = await db.students.find_one({"reg_no": data.reg_no}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Student already exists")
    
    valid_till = calculate_student_validity(data.current_year)
    student = Student(**data.model_dump(), valid_till=valid_till)
    await db.students.insert_one(student.model_dump())
    return student

@api_router.post("/students/bulk")
async def bulk_create_students(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    students_created = 0
    errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:  # Skip empty rows
                continue
            
            data = StudentCreate(
                reg_no=str(row[0]),
                name=str(row[1]),
                email=str(row[2]),
                mobile_no=str(row[3]),
                dob=str(row[4]),
                current_year=int(row[5])
            )
            
            # Check if exists
            existing = await db.students.find_one({"reg_no": data.reg_no}, {"_id": 0})
            if existing:
                errors.append(f"Row {idx}: Student {data.reg_no} already exists")
                continue
            
            valid_till = calculate_student_validity(data.current_year)
            student = Student(**data.model_dump(), valid_till=valid_till)
            await db.students.insert_one(student.model_dump())
            students_created += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    return {"success": True, "created": students_created, "errors": errors}

@api_router.get("/students", response_model=List[Student])
async def get_students():
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    return students

# ============ FACULTY ROUTES ============

@api_router.post("/faculty", response_model=Faculty)
async def create_faculty(data: FacultyCreate):
    faculty_id = await get_next_faculty_id()
    faculty = Faculty(**data.model_dump(), faculty_id=faculty_id)
    await db.faculty.insert_one(faculty.model_dump())
    return faculty

@api_router.post("/faculty/bulk")
async def bulk_create_faculty(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    faculty_created = 0
    errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:  # Skip empty rows
                continue
            
            data = FacultyCreate(
                name=str(row[0]),
                email=str(row[1]),
                mobile_no=str(row[2]),
                department=str(row[3]),
                profession=str(row[4]),
                valid_till=str(row[5])
            )
            
            faculty_id = await get_next_faculty_id()
            faculty = Faculty(**data.model_dump(), faculty_id=faculty_id)
            await db.faculty.insert_one(faculty.model_dump())
            faculty_created += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    return {"success": True, "created": faculty_created, "errors": errors}

@api_router.get("/faculty", response_model=List[Faculty])
async def get_faculty():
    faculty = await db.faculty.find({}, {"_id": 0}).to_list(1000)
    return faculty

# ============ EVENT ROUTES ============

@api_router.post("/events", response_model=Event)
async def create_event(data: EventCreate):
    event = Event(**data.model_dump())
    await db.events.insert_one(event.model_dump())
    return event

@api_router.get("/events", response_model=List[Event])
async def get_events():
    events = await db.events.find({}, {"_id": 0}).to_list(1000)
    return events

class EventStudentManualCreate(BaseModel):
    event_id: str
    reg_no: str
    name: str
    email: EmailStr
    mobile_no: str

@api_router.post("/events/students")
async def add_student_to_event(data: EventStudentCreate):
    # Get student details
    student = await db.students.find_one({"reg_no": data.reg_no}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in database. Use manual entry for external students.")
    
    # Get event details
    event = await db.events.find_one({"event_id": data.event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Create event student entry
    event_student = EventStudent(
        event_id=data.event_id,
        reg_no=student["reg_no"],
        name=student["name"],
        email=student["email"],
        mobile_no=student["mobile_no"],
        valid_from=event["date_from"],
        valid_to=event["date_to"]
    )
    
    await db.event_students.insert_one(event_student.model_dump())
    
    # Send QR email
    await send_qr_email(
        event_student.email,
        event_student.name,
        event_student.token,
        event_student.valid_to
    )
    
    return {"success": True, "message": "Student added to event and email sent"}

@api_router.post("/events/students/manual")
async def add_manual_student_to_event(data: EventStudentManualCreate):
    # Get event details
    event = await db.events.find_one({"event_id": data.event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Create event student entry with manual data
    event_student = EventStudent(
        event_id=data.event_id,
        reg_no=data.reg_no,
        name=data.name,
        email=data.email,
        mobile_no=data.mobile_no,
        valid_from=event["date_from"],
        valid_to=event["date_to"]
    )
    
    await db.event_students.insert_one(event_student.model_dump())
    
    # Send QR email
    await send_qr_email(
        event_student.email,
        event_student.name,
        event_student.token,
        event_student.valid_to
    )
    
    return {"success": True, "message": "External student added to event and email sent"}

@api_router.post("/events/students/bulk")
async def bulk_add_students_to_event(event_id: str = Form(...), file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")
    
    # Get event details
    event = await db.events.find_one({"event_id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    students_added = 0
    errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:  # Skip empty rows
                continue
            
            reg_no = str(row[0])
            
            # Check if it's a manual entry (has name, email, mobile in Excel)
            if len(row) >= 4 and row[1] and row[2] and row[3]:
                # Manual entry for external students
                event_student = EventStudent(
                    event_id=event_id,
                    reg_no=reg_no,
                    name=str(row[1]),
                    email=str(row[2]),
                    mobile_no=str(row[3]),
                    valid_from=event["date_from"],
                    valid_to=event["date_to"]
                )
            else:
                # Try to find in existing students
                student = await db.students.find_one({"reg_no": reg_no}, {"_id": 0})
                
                if not student:
                    errors.append(f"Row {idx}: Student {reg_no} not found and no manual data provided")
                    continue
                
                event_student = EventStudent(
                    event_id=event_id,
                    reg_no=student["reg_no"],
                    name=student["name"],
                    email=student["email"],
                    mobile_no=student["mobile_no"],
                    valid_from=event["date_from"],
                    valid_to=event["date_to"]
                )
            
            await db.event_students.insert_one(event_student.model_dump())
            
            # Send email
            await send_qr_email(
                event_student.email,
                event_student.name,
                event_student.token,
                event_student.valid_to
            )
            
            students_added += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    return {"success": True, "added": students_added, "errors": errors}

@api_router.get("/events/{event_id}/students", response_model=List[EventStudent])
async def get_event_students(event_id: str):
    students = await db.event_students.find({"event_id": event_id}, {"_id": 0}).to_list(1000)
    return students

# ============ VISITOR & ALUMNI ROUTES ============

@api_router.post("/visitors", response_model=Visitor)
async def register_visitor(data: Visitor):
    await db.visitors.insert_one(data.model_dump())
    
    # Send QR email
    await send_qr_email(data.email, data.name, data.token, data.valid_till)
    
    return data

@api_router.get("/visitors", response_model=List[Visitor])
async def get_visitors():
    visitors = await db.visitors.find({}, {"_id": 0}).to_list(1000)
    return visitors

@api_router.post("/alumni", response_model=Alumni)
async def register_alumni(data: Alumni):
    await db.alumni.insert_one(data.model_dump())
    
    # Send QR email
    await send_qr_email(data.email, data.name, data.token, data.valid_till)
    
    return data

@api_router.get("/alumni", response_model=List[Alumni])
async def get_alumni():
    alumni = await db.alumni.find({}, {"_id": 0}).to_list(1000)
    return alumni

# ============ QR VALIDATION ============

@api_router.post("/validate-qr")
async def validate_qr(data: QRValidate):
    token = data.token
    now = datetime.now(timezone.utc)
    
    # Check all collections
    collections = [
        ("students", "Student"),
        ("faculty", "Faculty"),
        ("visitors", "Visitor"),
        ("alumni", "Alumni"),
        ("event_students", "Event Student")
    ]
    
    for collection_name, user_type in collections:
        collection = db[collection_name]
        
        # Handle event_students with valid_to instead of valid_till
        if collection_name == "event_students":
            record = await collection.find_one({"token": token}, {"_id": 0})
            if record:
                valid_till_str = record.get("valid_to")
                valid_till = datetime.fromisoformat(valid_till_str.replace("Z", "+00:00"))
                
                if now <= valid_till:
                    return {
                        "valid": True,
                        "type": user_type,
                        "name": record.get("name"),
                        "email": record.get("email"),
                        "valid_till": valid_till_str,
                        "details": record
                    }
                else:
                    return {
                        "valid": False,
                        "reason": "QR code expired",
                        "type": user_type,
                        "name": record.get("name"),
                        "expired_on": valid_till_str
                    }
        else:
            record = await collection.find_one({"token": token}, {"_id": 0})
            if record:
                valid_till_str = record.get("valid_till")
                valid_till = datetime.fromisoformat(valid_till_str.replace("Z", "+00:00"))
                
                if now <= valid_till:
                    return {
                        "valid": True,
                        "type": user_type,
                        "name": record.get("name"),
                        "email": record.get("email"),
                        "valid_till": valid_till_str,
                        "details": record
                    }
                else:
                    return {
                        "valid": False,
                        "reason": "QR code expired",
                        "type": user_type,
                        "name": record.get("name"),
                        "expired_on": valid_till_str
                    }
    
    return {"valid": False, "reason": "QR code not found"}

# ============ DASHBOARD ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    
    # Count today's entries
    visitors_today = await db.visitors.count_documents({"created_at": {"$gte": today_start}})
    
    # Total counts
    total_students = await db.students.count_documents({})
    total_faculty = await db.faculty.count_documents({})
    total_visitors = await db.visitors.count_documents({})
    total_alumni = await db.alumni.count_documents({})
    total_events = await db.events.count_documents({})
    
    return {
        "visitors_today": visitors_today,
        "total_students": total_students,
        "total_faculty": total_faculty,
        "total_visitors": total_visitors,
        "total_alumni": total_alumni,
        "total_events": total_events
    }

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()